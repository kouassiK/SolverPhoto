import os
import re
import glob
import pathlib
import pandas as pd
import openpyxl 
import xlsxwriter
import collections
import datetime
import xml.etree.ElementTree as ET

tool_statuses = []

def FindSolutionPath():
    '''
    Fournit le chemin du fichier Excel dans lequel écrire les résultats
    '''
    path = os.getcwd()
    for filepath in pathlib.Path(path).glob('**/*'): 
        if(filepath.name == "result_presentation.xlsx" ): 
            f = filepath
    return f

def GetOrigin_Lots_Ord(name):
    '''
    Fournit horizon de référence pour calcul des lots ordonnancés
     =  origine(prise dans ToolStatuses.txt) + 4h
    '''
    global horizonrefs
    origin = ''
    path = os.getcwd()
    fichier = name.strip('.xml')[:-2] +'.txt'
    for filepath in pathlib.Path(path).glob('**/*'): 
        if(filepath.name == fichier ):
            f = open(filepath,'r')
            lines = f.readlines()
            line = lines[1] #Seconde ligne
            t = line.split(';') #séparation de la ligne selon les ;
            origin = t[4]# récupération de l'origine
            f.close()
    
    #print(fichier)
    date_str = origin.strip('\n')
    date_time_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')#formattage de la date
    t = date_time_obj.time()
    h = int(t.hour)
    m = int(t.minute)
    s = int(t.second)
    hours = datetime.time(h,0,0)
    a = datetime.timedelta(0, (h*3600+m*60+s))
    b = datetime.timedelta(0, (4*3600+0*60+0))
    horizonrefs = [a+b,datetime.time(h+4,m ,s )] #horizonsref en datetimedelta et datetime


def parse_lots_ordonnances(filename):
    '''
    Calcul des lots ordonnancés avant la date de début
    '''
    global m
    tree = ET.parse(filename)
    root = tree.getroot()
    keys = list(root[8][0].attrib.keys())
    tools = []
    d = []
    GetOrigin_Lots_Ord(filename)
    for i in range(len(root[8])):
        tools.append(root[8][i].get('ToolID'))
        d.append([root[8][i].get('ToolID'),root[8][i].get('StartDateTime')])
        
    #print(d)
    unique_tools_id = list(set(tools))
    data = dict([(key, []) for key in unique_tools_id])
    result = dict([(key, []) for key in unique_tools_id])
    occurrences = collections.Counter(tools)#contient le nombre d'occurences de chaque tool
    
    #print(occurrences)
    #print(unique_tools_id)
    for i in range(len(d)):
        if(d[i][0] in unique_tools_id):
            l = d[i][1].split("T")[1]
            #print(l)
            date_time_obj = datetime.datetime.strptime(l, '%H:%M:%S.%f')
            data[d[i][0]].append(date_time_obj.time())

    final_result = dict([(key, []) for key in unique_tools_id])
    for k in data.keys():
        l = data.get(k)
        a = len(data.get(k))
        print(l)
        for j in range(a):
            if(l[j] < horizonrefs[1]): #sélection des dates de début < horizon de référence
                result[k].append(l[j])


    #print(l)
    '''for k,v in result.items():
        print(k,v)'''


    print(horizonrefs)#horizon de ref
    '''for k,v in result.items():
        print(k,v)'''

    for k,v in result.items():
        final_result[k] = len(v)
        #print(k,v)
    
    '''for k,v in data.items():
        print(k,v)'''

    '''for k,v in final_result.items():
        print(k,v)'''
    return final_result


def GetOrigin_Lots_Taux(name, moves_windows):
    '''
    Fournit horizon de référence pour calcul du taux d'odonnancement 
     =  origine(prise dans ToolStatuses.txt) + moves_windows
    '''
    global refs
    origin = ''
    path = os.getcwd()
    fichier = name.strip('.xml')[:-2] +'.txt'
    for filepath in pathlib.Path(path).glob('**/*'): 
        if(filepath.name == fichier ):
            f = open(filepath,'r')
            lines = f.readlines()
            line = lines[1] #Get the second line
            t = line.split(';') #get table of value separated by semi colon
            origin = t[4]
            f.close()
    
    date_str = origin.strip('\n')
    date_time_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
    t = date_time_obj.time()
    h = int(t.hour)
    m = int(t.minute)
    s = int(t.second)
    hours = datetime.time(h,0,0)
    a = datetime.timedelta(0, (h*3600+m*60+s))
    b = datetime.timedelta(0, (moves_windows*3600+0*60+0))
    refs = [a+b,datetime.time(h+moves_windows,m ,s )] #horizonsref en datetimedelta et datetime


def parse_taux(filename,moves_windows): 
    mw = datetime.timedelta(0, (moves_windows*3600+0*60+0)) #conversion du moves windows pour les calculs
    '''
    Fourni le taux de fonctionnement des machines
    '''
    global m
    tree = ET.parse(filename)
    root = tree.getroot()
    keys = list(root[8][0].attrib.keys())
    tools = []
    d = []
    GetOrigin_Lots_Taux(filename, moves_windows) 
    for i in range(len(root[8])):
        tools.append(root[8][i].get('ToolID'))
        d.append([root[8][i].get('ToolID'),root[8][i].get('StartDateTime'),root[8][i].get('EndDateTime')])

    unique_tools_id = list(set(tools))
    data = dict([(key, []) for key in unique_tools_id])
    result = dict([(key, []) for key in unique_tools_id])
    final_result = dict([(key, []) for key in unique_tools_id])
    occurrences = collections.Counter(tools)#pour savoir la taille de la liste des dates selon l'id

    for i in range(len(d)):
        if(d[i][0] in unique_tools_id):
            l1 = d[i][1].split("T")[1]
            l2 = d[i][2].split("T")[1]
            #print(l)
            date_time_obj1 = datetime.datetime.strptime(l1, '%H:%M:%S.%f')
            date_time_obj2 = datetime.datetime.strptime(l2, '%H:%M:%S.%f')
            t1 = date_time_obj1.time()
            t2 = date_time_obj2.time()
            h1 = int(t1.hour)
            m1 = int(t1.minute)
            s1 = int(t1.second)
            
            h2 = int(t2.hour)
            m2 = int(t2.minute)
            s2 = int(t2.second)
            
            d1 = datetime.timedelta(0, (h1*3600+m1*60+s1))
            d2 = datetime.timedelta(0, (h2*3600+m2*60+s2))
            tableau = []
            tableau.append(d1)
            tableau.append(d2)
            data[d[i][0]].append(tableau)
                      
    #print('horizon de référence: ' , refs[1])
    for k in data.keys():
        debut = data.get(k)[0][0] #date origine 
        if(debut < refs[0] ): #date de début < horizon de réf
            fin  = datetime.timedelta(0, (0*3600+0*60+0))   #initialisation fin à 0   
            f = []
            for i in range(len(data.get(k))):   
                if(data.get(k)[i][0] < data.get(k)[i-1][1] and data.get(k)[i-1][1] < refs[0] ):                 
                    fin = data.get(k)[i][1]
            
                if(data.get(k)[i][0] > data.get(k)[i-1][1] and data.get(k)[i-1][1] < refs[0]):
                    t = []
                    tbis = []
                    debut_prec = data.get(k)[i-1][0]
                    fin_prec = data.get(k)[i-1][1]
                    t.append(debut_prec)
                    t.append(fin_prec) 
                    debut = data.get(k)[i][0]
                    fin = data.get(k)[i][1]
                    result[k].append(t)
                    
            f.append(debut)
            f.append(fin)
            result[k].append(f)
                
    taux = dict([(key, 0) for key in unique_tools_id])
    for k in result.keys():
        s = datetime.timedelta(0, (0*3600+0*60+0))
        for i in range(len(result.get(k))):
            if(result.get(k)[i][1] > result.get(k)[i][0]):
                s = s + result.get(k)[i][1] - result.get(k)[i][0]
            if(result.get(k)[i][1] < result.get(k)[i][0]):
                s = datetime.timedelta(seconds = (result.get(k)[i][1] - result.get(k)[i][0]).seconds/3600)
        taux[k] = (s /  mw)*100 #taux en pourcentage

    '''for k,v in taux.items():
        print(k,v)'''
    return taux  

def WriteResults():
    files = []
    path = os.getcwd()
    #On récupère tous les noms des extracts dans files
    for f in glob.glob("*.xlsx"):
        if(f != 'result_presentation.xlsx'):
            files.append(f.strip('.xlsx'))
            
    for i in range(len(files)): 
        for filepath in pathlib.Path(path).glob('**/*'): 
            if(filepath.name.find(files[i]+ '2h.xml')!= -1):
                o1 = parse_lots_ordonnances(files[i] + '2h.xml')
                t1 = parse_taux(files[i] + '2h.xml',2)
            if(filepath.name.find(files[i]+ '3h.xml')!= -1):
                o2 = parse_lots_ordonnances(files[i] + '3h.xml') 
                t2 = parse_taux(files[i] + '3h.xml',3)
            if(filepath.name.find(files[i]+ '4h.xml')!= -1):
                o3 = parse_lots_ordonnances(files[i] + '4h.xml')
                t3 = parse_taux(files[i] + '4h.xml', 4) 
            if(filepath.name.find(files[i]+ '5h.xml')!= -1):
                o4 = parse_lots_ordonnances(files[i] + '5h.xml')
                t4 = parse_taux(files[i] + '5h.xml', 5)
            if(filepath.name.find(files[i]+ '6h.xml')!= -1):
                o5 = parse_lots_ordonnances(files[i] + '6h.xml')
                t5 = parse_taux(files[i] + '6h.xml', 6)
        
        Fill(files[i],o1,o2,o3,o4,o5,t1,t2,t3,t4,t5)  
            


def Fill(extract,
o1,o2,o3,o4,o5,
t1,t2,t3,t4,t5):
    '''
    Sauvegarde des résultats dans result_presentation.xlsx
    '''
    global lignes
    lignes,tools, ids, sheets = [], [], [], []
    path = os.getcwd() 
    #print(path)
    for filepath in pathlib.Path(path).glob('**/*'): 
        #print(filepath.name)
        if(filepath.name == "Processability_"+ extract +".txt"):
            f = open(filepath,'r')
            lines = f.readlines()
            lines.remove(lines[0])
            for line in lines:
                line.strip('\n')
                list1 = line.split(';')
                #print(list1)
                ids.append(list1[2])
                #liste des outils sans doublons
            f.close()

    machines = list(set(ids))
    #print(machines)
    occurrences = collections.Counter(ids)
    workbook = openpyxl.load_workbook('result_presentation.xlsx') 
    sheets = workbook.sheetnames
    i = sheets.index(extract)
    worksheet = workbook.worksheets[i]
    row = worksheet.max_row
    j = 3
    #on écrit le nom des tools dans le fichier Excel
    for i in range(1,len(machines)+1):
        worksheet.cell(j , 1).value = machines[i-1]
        worksheet.cell(j , 2).value = occurrences[machines[i-1]]
        worksheet.cell(j , 3).value = o1.get(machines[i-1])
        worksheet.cell(j , 4).value = o2.get(machines[i-1])
        worksheet.cell(j , 5).value = o3.get(machines[i-1])
        worksheet.cell(j , 6).value = o4.get(machines[i-1])
        worksheet.cell(j , 7).value = o5.get(machines[i-1])
        worksheet.cell(j , 8).value = t1.get(machines[i-1])
        worksheet.cell(j , 9).value = t2.get(machines[i-1])
        worksheet.cell(j , 10).value = t3.get(machines[i-1])
        worksheet.cell(j , 11).value = t4.get(machines[i-1])
        worksheet.cell(j , 12).value = t5.get(machines[i-1])
        j = j + 1
    workbook.save('result_presentation.xlsx')



WriteResults()



