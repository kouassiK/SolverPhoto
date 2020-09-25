import os
import glob
import pathlib
import openpyxl 
import xlsxwriter
import collections
import datetime
import dateutil.parser
import statistics
from collections import OrderedDict
import xml.etree.ElementTree as ET
import math
from orderedset import OrderedSet

#la liste complète des lots de Lots.txt
def liste_lots_fab():
    name = "Lots.txt"
    path = os.getcwd()
    lines = []
    lots = []
    for filepath in pathlib.Path(path).glob('**/*'): 
        if(filepath.name == name ):
            f = open(filepath,'r')
            c = len(open(filepath).readlines( ))
    for i in range(c):
        line = f.readline()
        t = line.split(';')
        lines.append(t)

    lots = [] #Dans Lots.txt
    
    for i in range(1,len(lines)):
        lots.append(lines[i][0])

    #print(len(lots))
    return lots
    
#la liste des lots qui figurent dans la solution du solveur pendant l'horizon de planification
def liste_lots_solveur(nameXml):
    tree = ET.parse(nameXml)
    root = tree.getroot()
    infos = []
    for i in range(len(root[4])):
        infos.append(root[4][i].get('LotID'))
    #print(len(infos))
    return infos

#la liste des lots qui figurent dans la (toute)solution du solveur hors et dans l' horizon de planification
def liste_lots_solveur_complete(nameXml):
    tree = ET.parse(nameXml)
    root = tree.getroot()
    masques = []
    infos = []
    
    for i in range(len(root[8])):
        masques.append(root[8][i].get('MaskID'))  
        date = dateutil.parser.isoparse(root[8][i].get('OperationStartDateTime'))
        debut_traitement = dateutil.parser.isoparse(root[8][i].get('StartDateTime'))      
        infos.append([root[8][i].get('LotID'), root[8][i].get('MaskID'), date, root[8][i].get('PriorityClass'), root[8][i].get('RouteStepID') , debut_traitement ]) #lotId, masque, operStart, priority,routestepid, debut_traitement
    
    data = dict([(key, []) for key in masques])

    for i in range(len(masques)):
        if(masques[i] in data.keys()): #si l'id lot est une clé
            data[masques[i]].append(infos[i])

    '''for k,v in data.items():
        print(k,v)'''

    #print(len(data))

    return data


#dictionnaire de lotId, masque, operStart, priority,routestepid avec données de Lots.txt avec clé maskId (fab)
def lots_fab():
    name = "Lots.txt"
    path = os.getcwd()
    lines = []
    masks = []
    for filepath in pathlib.Path(path).glob('**/*'): 
        if(filepath.name == name ):
            f = open(filepath,'r')
            c = len(open(filepath).readlines( ))
    for i in range(c):
        line = f.readline()
        t = line.split(';')
        lines.append(t)

    lots = [] #Dans Lots.txt
    
    for i in range(1,len(lines)):
        masks.append(lines[i][21]) 
        date = dateutil.parser.isoparse(lines[i][6].strip('\n'))
        lots.append([lines[i][0], lines[i][21], date, str(lines[i][5]),lines[i][1] ] ) #lotId, masque, operStart, priority,routestepid

    masks_ids = set(masks)
    data_lots = dict([(key, []) for key in masks_ids])


    for i in range(len(lots)):
        if(lots[i][1] in data_lots.keys()): #si le masque est une clé
            data_lots[lots[i][1]].append(lots[i])

    '''for k,v in data_lots.items():
        print(k,v)'''

    return data_lots

#dictionnaire lotID, MaskID, date, priority, routestepid, date début traitement des lots de la solution solveur avec clé maskId (solveur)
def proposition_solveur(nameXml):
    tree = ET.parse(nameXml)
    root = tree.getroot()
    infos = []
    masks = []
    for i in range(len(root[4])):
        masks.append(root[4][i].get('MaskID'))
        date = dateutil.parser.isoparse(root[4][i].get('OperationStartDateTime'))
        debut_traitement = dateutil.parser.isoparse(root[4][i].get('StartDateTime'))
        infos.append([root[4][i].get('LotID'), root[4][i].get('MaskID'), date , root[4][i].get('PriorityClass'),root[4][i].get('RouteStepID'), debut_traitement]) #ids des lots de la solution solveur
 
    m = set(masks)
    d = dict([(key, []) for key in m])
    for i in range(len(infos)):
        if(infos[i][1] in d.keys()): #si le masque est une clé
            d[infos[i][1]].append(infos[i])

    '''for k,v in d.items():
        print(k,v)'''

    return d


#dictionnaire lots dont le masque et la routestepid sont dans la solution mais qui n'est pas proposé dans la solution
def lots_oubli(lots_fab, lots_solveur, dict_fab, dict_solveur):
    l = list(set(lots_fab).symmetric_difference(set(lots_solveur))) #lots dans fab mais pas dans la solution solveur
    masks = [] 
    routes = []
    #print(l)

    '''for k,v  in dict_fab.items(): #tous les lots et informations (dans et hors solution)
        print(k,v)'''

    '''for k,v  in dict_solveur.items(): #tous les lots et informations (dans la solution)
        print(k,v)'''

    #print(len(dict_fab), len(dict_solveur))

    for k in dict_solveur.keys():
        for i in range(len(dict_solveur.get(k))):
            routes.append(dict_solveur.get(k)[i][4]) #les recipe_group

    routes_solveur = set(routes) #les recipe_group utilisés pendant 2h
    #print(len(routes_solveur))

    for k in dict_fab.keys():
        for i in range(len(dict_fab.get(k))):
            #print(dict_fab.get(k)[i][0])
            if((dict_fab.get(k)[i][0] in l) and (k  in dict_solveur.keys())) :
                masks.append(k) #masques dans le solveur pour lesquels il existe des lots oubliés

    #print(masks)

    result_comparaison = dict([(key, []) for key in masks]) 
    for k in dict_fab.keys():
        for i in range(len(dict_fab.get(k))):
                if((dict_fab.get(k)[i][0] in l) and (k  in dict_solveur.keys()) and (dict_fab.get(k)[i][4] in routes_solveur)): #dans la fab et pas réalisé 
                    #print(dict_fab.get(k)[i][0])
                    result_comparaison[k].append(dict_fab.get(k)[i])
       
    '''for k,v in result_comparaison.items():
        print(k,v)''' #lots ne figurant pas dans la solution mais dont le masque et la recette sont utilisés

    return result_comparaison


#liste_lots_solveur('[0.001_1000_0.001](8_2_1).xml')
#lots_oubli(liste_lots_fab(), liste_lots_solveur('[0_1_0](8_2_1).xml'), liste_lots_solveur_complete('[0_1_0](8_2_1).xml'), proposition_solveur('[0_1_0](8_2_1).xml'))
#differences('1', '[0_1_0](8_2_1).xml', liste_lots_fab(), liste_lots_solveur('[0_1_0](8_2_1).xml'), lots_fab(), proposition_solveur('[0_1_0](8_2_1).xml'))

#dictionnaire des dates par ordre chronologique , clé maskID. Il s'agit des masques dans la solution pour lesquels certains lots ne sont pas encore placés
def solveur_hours(nameXml, lots_fab, lots_solveur, dict_fab, dict_solveur):
    masques_utilises_lots_oubli = lots_oubli(lots_fab, lots_solveur, dict_fab, dict_solveur).keys() #masques utilisés
    solver = proposition_solveur(nameXml)   #dictionnaire de la solution solveur
    resultat = dict((key,0) for key in masques_utilises_lots_oubli)

    for k in solver.keys():
        dates_ordre = []
        if(k in masques_utilises_lots_oubli):
            liste_dates = solver.get(k)
            dates_ordre = sorted(liste_dates)  #dates dans l'ordre chronologique
            #print(dates_ordre)
        for i in range(len(dates_ordre)):
            resultat[k] = dates_ordre

    '''for k,v in resultat.items():
        print(k, v)'''
    return resultat

#Création fichiers Excel
def Excel():
    workbook = xlsxwriter.Workbook('Resultats_Propositions.xlsx')
    workbook.close()
    workbook1 = xlsxwriter.Workbook('Valeurs_Moyennes.xlsx')
    workbook1.close()


def differences(index,nameXml,lots_fab, lots_solveur, data_fab, data_solveur):
    dates_non_proposees = lots_oubli(lots_fab, lots_solveur, data_fab, data_solveur)
    dates_proposees = solveur_hours(nameXml,lots_fab, lots_solveur, data_fab, data_solveur) #en ordre chronologique
    cles_lots = []

    res = dict([(key, []) for key in dates_non_proposees.keys()])
    '''for k,v in dates_non_proposees.items():
        print(k,v)'''
    
    for k in dates_non_proposees.keys():
        for i in range(len(dates_non_proposees.get(k))):
            for j in range(len(dates_proposees.get(k))):
                l = []
                if(dates_non_proposees.get(k)[i][2] < dates_proposees.get(k)[j][2]): #comparaison des dates d'arrivée en atelier
                    l = [dates_non_proposees.get(k)[i][0], dates_non_proposees.get(k)[i][2], dates_non_proposees.get(k)[i][3],dates_non_proposees.get(k)[i][4], dates_non_proposees.get(k)[i][5]
                    ,dates_proposees.get(k)[j][0], dates_proposees.get(k)[j][2], dates_proposees.get(k)[j][3],dates_proposees.get(k)[j][4], dates_proposees.get(k)[j][5]]
                    res[k].append(l)
                    cles_lots.append(dates_non_proposees.get(k)[i][0])

        #print(l)
    
    cles = []
    for k in res.keys():
        if(len(res.get(k)) > 0):
            cles.append(k)

    final = dict([(key, []) for key in cles])

    for k in final.keys():
        final[k] = res.get(k)
    
    #print(cles_lots)
    value  = dict([(key, []) for key in cles_lots])
    l = OrderedDict([(key, []) for key in cles_lots]) #pour éliminer les doublons en évitant de mélanger des choses

    for k,v in final.items():
        print(k,v)

    for k in final.keys(): 
        for i in range(len(final.get(k))):    
            if(final.get(k)[i][0] in value.keys()):
                l[final.get(k)[i][0]].append(k)
                l[final.get(k)[i][0]].append(final.get(k)[i][0])#nom lot non proposé
                l[final.get(k)[i][0]].append(final.get(k)[i][1])#date entrée en fab non proposé
                l[final.get(k)[i][0]].append(final.get(k)[i][2])#priorité non proposé
                l[final.get(k)[i][0]].append(final.get(k)[i][3])#route non proposé
                l[final.get(k)[i][0]].append(final.get(k)[i][4])#date début de traitement non proposé

                value[final.get(k)[i][0]].append(final.get(k)[i][5])# nom proposé
                value[final.get(k)[i][0]].append((final.get(k)[i][6] - final.get(k)[i][1])) #écart temporel proposé par rapport à non proposé
                value[final.get(k)[i][0]].append(final.get(k)[i][7]) #priorité proposé
                value[final.get(k)[i][0]].append(final.get(k)[i][4]  - final.get(k)[i][9]) #date début de traitement proposé - date début de traitement non proposé
                
               
                final.get(k)[i][4] - final.get(k)[i][1] #début traitement  - début entrée en fab () -> non proposé
                final.get(k)[i][9]


    #for k in l.keys():
        #print(OrderedSet(l.get(k)))
    
    workbook = openpyxl.load_workbook('Resultats_Propositions.xlsx') 
    worksheet = workbook.create_sheet(index)
    
    worksheet.cell(1,1).value = nameXml
    worksheet.cell(2,1).value =  "FAB"
    worksheet.cell(3,1).value =  "Masks"
    worksheet.cell(3,2).value =  "Lots"
    worksheet.cell(3,3).value =  "Operation"
    worksheet.cell(3,4).value =  "Priority"
    worksheet.cell(3,5).value =  "RouteStep"
   

    worksheet.cell(2,8).value =  "SOLVEUR"
    worksheet.cell(3,8).value =  "Lots"
    worksheet.cell(3,9).value =  "Diff_Entree"
    worksheet.cell(3,10).value =  "Priority"
    #worksheet.cell(3,9).value =  "RouteStep"
    worksheet.cell(3,11).value =  "Diff_Traitement"

    
    ligne = 4
    for k in value.keys():
        for i in range(len(value.get(k))):
            worksheet.cell(ligne, 1).value = OrderedSet(l.get(k))[0]   #masque
            worksheet.cell(ligne, 2).value = k                          #lot
            worksheet.cell(ligne, 3).value = OrderedSet(l.get(k))[2] 
            worksheet.cell(ligne, 4).value = OrderedSet(l.get(k))[3] #priorité
            worksheet.cell(ligne, 5).value = OrderedSet(l.get(k))[4]  #route
            worksheet.cell(ligne, i + 8).value = value.get(k)[i] #info
        ligne = ligne + 1

    workbook.save('Resultats_Propositions.xlsx')


#differences('1', '[0_1_0](8_2_1).xml', liste_lots_fab(), liste_lots_solveur('[0_1_0](8_2_1).xml'), liste_lots_solveur_complete('[0_1_0](8_2_1).xml'), proposition_solveur('[0_1_0](8_2_1).xml'))

#Comparer les dates des lots proposés (par le solveur) par masque 
def comparaison(index,nameXml,lots_fab, lots_solveur, data_fab, data_solveur):
    workbook = openpyxl.load_workbook('Resultats_Propositions.xlsx') 
    worksheet = workbook.create_sheet(index)
    
    worksheet.cell(1,1).value = nameXml
    worksheet.cell(2,1).value =  "FAB"
    worksheet.cell(3,1).value =  "Masks"
    worksheet.cell(3,2).value =  "Lots"
    worksheet.cell(3,3).value =  "Operation"
    worksheet.cell(3,4).value =  "Priority"

    worksheet.cell(2,6).value =  "SOLVEUR"
    worksheet.cell(3,6).value =  "Lots"
    worksheet.cell(3,7).value =  "OperDate"
    worksheet.cell(3,8).value =  "Priority"

    dates_non_proposees = lots_oubli(lots_fab, lots_solveur, data_fab, data_solveur)
    dates_proposees = solveur_hours(nameXml,lots_fab, lots_solveur, data_fab, data_solveur) #en ordre chronologique

    final = dict([(key,[]) for key in dates_non_proposees.keys()])

    ligne = 4
    for k in dates_non_proposees.keys():
        for i in range(len(dates_non_proposees.get(k))):
            for j in range(len(dates_proposees.get(k))):
                if( dates_non_proposees.get(k)[i][1] < dates_proposees.get(k)[j][1] ):
                    worksheet.cell(ligne, 1).value = k                               #mask
                    worksheet.cell(ligne, 2).value = dates_non_proposees.get(k)[i][0] #lot
                    worksheet.cell(ligne, 3).value = dates_non_proposees.get(k)[i][1] #date
                    worksheet.cell(ligne, 4).value = dates_non_proposees.get(k)[i][2] #priorité

                    worksheet.cell(ligne, j + 6 ).value = dates_proposees.get(k)[j][1] - dates_non_proposees.get(k)[i][1] #différence des dates
                    worksheet.cell(ligne, j + 7 ).value = dates_proposees.get(k)[j][2] #priorité


            ligne = ligne + 1
    workbook.save('Resultats_Propositions.xlsx')


def ExistsXml():
    filesnames = []
    path = os.getcwd()
    for filepath in pathlib.Path(path).glob('**/*'): 
        if(filepath.name.find('.xml')!=-1 and filepath.name.find('[')!= -1 ) :
            filesnames.append(filepath.name) 
    return filesnames




Excel() # Création fichier Excel
fichiers = ExistsXml()
for i in range(len(fichiers)):
    differences(str(i), fichiers[i], liste_lots_fab(), liste_lots_solveur(fichiers[i]), liste_lots_solveur_complete(fichiers[i]), proposition_solveur(fichiers[i]))



#differences(str('1'), '[0_1_0](8_2_1).xml', liste_lots_fab(), liste_lots_solveur('[0_1_0](8_2_1).xml'), liste_lots_solveur_complete('[0_1_0](8_2_1).xml'), proposition_solveur('[0_1_0](8_2_1).xml')) 
#liste_lots_solveur_complete('[0.001_1000_0.001](8_2_1).xml')
#lots_oubli(liste_lots_fab(), liste_lots_solveur('[0.001_1000_0.001](8_2_1).xml'), lots_fab(), proposition_solveur('[0.001_1000_0.001](8_2_1).xml'))'''