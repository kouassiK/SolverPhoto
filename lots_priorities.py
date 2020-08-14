import os
import glob
import pathlib
import pandas as pd
import openpyxl 
import xlsxwriter
import datetime
import dateutil.parser
import xml.etree.ElementTree as ET


def parse_lots(name):
    #name = nom instance
    '''
    All LotsID and Priority taken from Lots.txt of each data 
    '''
    path = os.getcwd()
    lines = []
    c = 0
    for filepath in pathlib.Path(path).glob('**/*'): 
        if(filepath.name == name ):
            f = open(filepath,'r')
            c = len(open(filepath).readlines( ))
    for i in range(c):
        line = f.readline()
        t = line.split(';')
        lines.append(t)
    #print(lines)
    data = []
    for i in range(1,len(lines)):
        lotID = lines[i][0]
        priority = lines[i][5]
        data.append([lotID, priority])

    #print(data)
    return data

def total_high(name):
    '''
    Total number of high priority lots in Lots.txt
    '''
    data = parse_lots(name)
    high = 0
    for i in range(len(data)):
        if(data[i][1]) == 'High_PIT':
                high = high + 1
    print('high: ' ,high)
    return high

def total_standard(name):
    '''
    Total number of standard priority lots in Lots.txt
    '''
    data = parse_lots(name)
    standard = 0
    for i in range(len(data)):
        if(data[i][1]) == 'Standard_PIT':
                standard = standard + 1
    print('standard:' , standard)
    return standard

def total_customer(name):
    '''
    Total number of customer priority lots in Lots.txt
    '''
    data = parse_lots(name)
    custom = 0
    for i in range(len(data)):
            if(data[i][1]) == 'Customer_PIT':
                custom = custom + 1
    print('customer:' ,custom)
    return  custom


def parse_solution(nameXml):
    tree = ET.parse(nameXml)
    root = tree.getroot()
    lotsID_list = [] 
    for i in range(len(root[4])):
        lotsID_list.append([root[4][i].get('LotID'), root[4][i].get('StartDateTime'), root[4][i].get('EndDateTime')] )
    #print(lotsID_list)
    return lotsID_list

def total_priorities(name, high, standard, customer):
    #Pris sans le fichier Lots.txt
    fichier = open(name, "w") 
    fichier.write('Lots.txt' +'\n') 
    fichier.write('total high : ' +str(high) +'\n')
    fichier.write('total standard: ' + str(standard) +'\n')
    fichier.write('total customer: ' + str(customer) + '\n')
    fichier.close() 

def total_priorities_solution(nameXml, data):
    name = nameXml.split('h')[0] +'h' +'.txt'
    #print(name)
    fichier = open(name, "a") 
    fichier.write(nameXml +'\n')
    fichier.write('total high : ' + str(data.get('HIGH_PIT')) +'\n')
    fichier.write('total standard : ' + str(data.get('STANDARD_PIT')) +'\n')
    fichier.write('total customer : ' + str(data.get('CUSTOMER_PIT')) +'\n')

#Pour un seul set de paramètres
def percentages(nameXml, name):
    data_lots = parse_lots(name)
    data_solution = []
    for i in range(len(parse_solution(nameXml))):
        data_solution.append(parse_solution(nameXml)[i][0])

    HIGH_PIT = 0
    STANDARD_PIT = 0
    CUSTOMER_PIT = 0

    high_duration = datetime.timedelta(0, (0*3600+0*60+0))
    standard_duration = datetime.timedelta(0, (0*3600+0*60+0))
    customer_duration = datetime.timedelta(0, (0*3600+0*60+0))
    
    for i in range(len(data_lots)):
        if(data_lots[i][0] in data_solution):
            j = data_solution.index(data_lots[i][0]) #récupération de l'indice du lot qui est dans Lots.txt et la Solution.xml

            if(data_lots[i][1]) == 'High_PIT':
                print(j)
                HIGH_PIT = HIGH_PIT + 1
                high_duration = high_duration + dateutil.parser.isoparse(parse_solution(nameXml)[j][2])  - dateutil.parser.isoparse(parse_solution(nameXml)[j][1]) 
                #print(data_lots[i][0])

            if(data_lots[i][1]) == 'Standard_PIT':
                STANDARD_PIT = STANDARD_PIT + 1
                standard_duration = standard_duration + dateutil.parser.isoparse(parse_solution(nameXml)[j][2])  - dateutil.parser.isoparse(parse_solution(nameXml)[j][1])
            
            if(data_lots[i][1]) == 'Customer_PIT':
                CUSTOMER_PIT = CUSTOMER_PIT + 1
                customer_duration = customer_duration + dateutil.parser.isoparse(parse_solution(nameXml)[j][2])  - dateutil.parser.isoparse(parse_solution(nameXml)[j][1])
    #print( HIGH_PIT, STANDARD_PIT,CUSTOMER_PIT )

    final = {'HIGH_PIT' : [100*(HIGH_PIT / total_high(name)), high_duration.total_seconds() / 3600],  'STANDARD_PIT': [100*(STANDARD_PIT / total_standard(name)), standard_duration.total_seconds() / 3600], 'CUSTOMER_PIT' : [100*(CUSTOMER_PIT / total_customer(name)), customer_duration.total_seconds() / 3600 ]}
    
    for(k,v) in final.items():
        print(k,v)

    return final


def ExcelTemplate():      
    '''
    Excel template to write the results
    '''
    workbook = xlsxwriter.Workbook('FinalResults_0_1_0.xlsx') 
    worksheet = workbook.add_worksheet() 
    worksheet.write('A1', 'Parameters ')
    worksheet.write('B1', 'High priority % ') 
    worksheet.write('C1', 'Standard priority % ') 
    worksheet.write('D1', 'Customer priority %')
    worksheet.write('E1', 'High hours ') 
    worksheet.write('F1', 'Standard hours') 
    worksheet.write('G1', 'Customer hours ')

    worksheet.write(1, 0, "(0,0,0)") #1
    worksheet.write(2, 0, "(1,0,0)") #2
    worksheet.write(3, 0, "(2,1,0.5)" ) #3
    worksheet.write(4, 0, "(3,1,0.5)") #4
    worksheet.write(5, 0,"(8,2,1)" ) #5
    worksheet.write(6, 0, "(10,4,2)") #6
    worksheet.write(7, 0, "(20,2,1)") #7
    worksheet.write(8, 0,"(100,0.1,0.01)" ) #8
    worksheet.write(9, 0,"(1000,20,1)" ) #9

    workbook.close()


def Edit(longname, text_name):
    '''
    Saving the results in the Excel file
    '''
    name = longname.split('h_')[0]
    parameters = (longname.strip('.xml')).split('h_')[1]
    #print(parameters)
    i = 0
    workbook = openpyxl.load_workbook('FinalResults_0_1_0.xlsx') 

    l = workbook.sheetnames
    if name in l: # Excel existe et la feuille existe déjà
        worksheet = workbook[name]
        #on trouve sa ligne
        if(parameters.find("0_0_0")!= -1): 
            i = 2
        if(parameters.find("1_0_0")!= -1): 
            i = 3
        if(parameters.find("2_1_0.5")!= -1):
            i = 4
        if(parameters.find("3_1_0.5")!= -1):
            i = 5      
        if(parameters.find("8_2_1")!= -1):
            i = 6
        if(parameters.find("10_4_2")!= -1):
            i = 7
        if(parameters.find("20_2_1")!= -1):
            i = 8
        if(parameters.find("100_0.1_0.01")!= -1):
            i = 9
        if(parameters.find("1000_20_1")!= -1):
            i = 10
        

        worksheet.cell(i,2).value = percentages(longname, text_name).get('HIGH_PIT')[0]
        worksheet.cell(i,3).value = percentages(longname, text_name).get('STANDARD_PIT')[0]
        worksheet.cell(i,4).value = percentages(longname, text_name).get('CUSTOMER_PIT')[0]
        worksheet.cell(i,5).value = percentages(longname, text_name).get('HIGH_PIT')[1]
        worksheet.cell(i,6).value = percentages(longname, text_name).get('STANDARD_PIT')[1]
        worksheet.cell(i,7).value = percentages(longname, text_name).get('CUSTOMER_PIT')[1]

    workbook.save("FinalResults_0_1_0.xlsx")
    if name not in l: # Excel existe et la feuille n'existe pas 
        #on crée la feuille
        worksheet = workbook.create_sheet(name)
        worksheet.cell(1,1).value =  'Parameters '
        worksheet.cell(1,2).value = 'High priority % '
        worksheet.cell(1,3).value = 'Standard priority % ' 
        worksheet.cell(1,4).value =  'Customer priority %'

        worksheet.cell (1,6).value = 'High hours '
        worksheet.cell (1,7).value   = 'Standard hours' 
        worksheet.cell (1,8).value = 'Customer hours'

        worksheet.cell(2, 1).value =  "(0,0,0)" #1
        worksheet.cell(3, 1).value =  "(1,0,0)" #2
        worksheet.cell(4, 1).value =  "(2,1,0.5)"  #3
        worksheet.cell(5, 1).value =  "(3,1,0.5)"  #4
        worksheet.cell(6, 1).value = "(8,2,1)"  #5
        worksheet.cell(7, 1).value = "(10,4,2)" #6
        worksheet.cell(8, 1).value = "(20,2,1)"  #7
        worksheet.cell(9, 1).value = "(100,0.1,0.01)"  #7
        worksheet.cell(10, 1).value = "(1000,20,1)"  #8
        

        #on trouve sa ligne
        if(parameters.find("0_0_0")!= -1): 
            i = 2
        if(parameters.find("1_0_0")!= -1): 
            i = 3
        if(parameters.find("2_1_0.5")!= -1):
            i = 4
        if(parameters.find("3_1_0.5")!= -1):
            i = 5
        if(parameters.find("8_2_1")!= -1):
            i = 6
        if(parameters.find("10_4_2")!= -1):
            i = 7
        if(parameters.find("20_2_1")!= -1):
            i = 8
        if(parameters.find("100_0.1_0.01")!= -1):
            i = 9 
        if(parameters.find("1000_20_1")!= -1):
            i = 10   
        
    #on remplit
        worksheet.cell(i,2).value = percentages(longname, text_name).get('HIGH_PIT')[0]
        worksheet.cell(i,3).value = percentages(longname, text_name).get('STANDARD_PIT')[0]
        worksheet.cell(i,4).value = percentages(longname, text_name).get('CUSTOMER_PIT')[0]
        worksheet.cell(i,5).value = percentages(longname, text_name).get('HIGH_PIT')[1]
        worksheet.cell(i,6).value = percentages(longname, text_name).get('STANDARD_PIT')[1]
        worksheet.cell(i,7).value = percentages(longname, text_name).get('CUSTOMER_PIT')[1]

    workbook.save('FinalResults_0_1_0.xlsx')



def ExistsXlsx():
    '''
    Check if the Excel file containing the results exists
    '''
    filesnames = []
    path = os.getcwd()
    value = False
    for filepath in pathlib.Path(path).glob('**/*'): 
        if(filepath.name == "FinalResults_0_1_0.xlsx"):
            value = True
    return value 

def ExistsXml():
    '''
    Get the list of all xml files
    '''
    filesnames = []
    path = os.getcwd()
    for filepath in pathlib.Path(path).glob('**/*'): 
        if(filepath.name.find('.xml')!=-1 and filepath.name.find('Solver')!=-1):
            filesnames.append(filepath.name) 
    return filesnames


def SolverExcel():
    '''
    Get the list of all text files with Solver in the name
    '''
    excelfiles = []
    path = os.getcwd()
    for filepath in pathlib.Path(path).glob('**/*'): 
        if(filepath.name.find('.xlsx')!=-1 and filepath.name.find('Solver')!=-1):
            excelfiles.append(filepath.name) 
    return excelfiles

def LotsTxt():
     '''
    Get the list of all text files with Lots in the name
    '''
    txtfiles = []
    path = os.getcwd()
    for filepath in pathlib.Path(path).glob('**/*'): 
        if(filepath.name.find('.txt')!=-1 and filepath.name.find('Lots')!=-1):
            txtfiles.append(filepath.name) 
    return txtfiles

'''total_high('Lots_Solver_20200710_10h.txt')
total_standard('Lots_Solver_20200710_10h.txt')
total_customer('Lots_Solver_20200710_10h.txt')'''
#parse_lots('Lots_Solver_20200710_10h.txt')
#Edit(fichiers[i], 'Lots_' + fichiers[i].split('h')[0] + 'h' + '.txt')


#percentages('Solver_20200721_15h2h_100_0.1_0.01.xml', 'Lots_Solver_20200721_15h.txt')

ExcelTemplate()
fichiers = ExistsXml()
for i in range(len(fichiers)):
    #total_priorities_solution(fichiers[i], percentages(fichiers[i], 'Lots_' +fichiers[i].split('h')[0] + 'h' + '.txt'))
    Edit(fichiers[i], 'Lots_' + fichiers[i].split('h')[0] + 'h' + '.txt')

'''lots = LotsTxt()
for i in range(len(lots)):
    name = lots[i] 
    total_priorities(name.strip('Lots_txt.') +'.txt', total_high(name), total_standard(name), total_customer(name))'''
